import os
import json
from datetime import datetime
import openpyxl

class BatchManager:
    BATCHES_PATH = os.path.join('results', 'batches.json')

    def __init__(self):
        self.batches = {}
        self.current_batch = 'Default'
        self.load()

    def load(self):
        if os.path.exists(self.BATCHES_PATH):
            with open(self.BATCHES_PATH, 'r') as f:
                self.batches = json.load(f)
        else:
            self.batches = {'Default': []}

    def save(self):
        os.makedirs(os.path.dirname(self.BATCHES_PATH), exist_ok=True)
        with open(self.BATCHES_PATH, 'w') as f:
            json.dump(self.batches, f, indent=2)

    def create_batch(self, name=None):
        if not name:
            name = f"Batch_{datetime.now().strftime('%Y%m%d_%H%M')}"
        self.batches[name] = []
        self.current_batch = name
        self.save()
        return name

    def delete_batch(self, name):
        if name in self.batches and name != 'Default':
            del self.batches[name]
            self.save()

    def get_excel_path(self, batch_name=None):
        if not batch_name:
            batch_name = self.current_batch
        safe_name = batch_name.replace(' ', '_')
        return os.path.join('results', f'batch_{safe_name}.xlsx')

    def add_result(self, result):
        if self.current_batch not in self.batches:
            self.batches[self.current_batch] = []
        self.batches[self.current_batch].append(result)
        self.save()
        # Excel export logic
        excel_path = self.get_excel_path()
        from openpyxl import Workbook, load_workbook
        import os
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = 'Results'
            # Write header
            ws.append([
                'Board Number', 'Timestamp', 'Result', 'Missing', 'Detected', 'Expected', 'Batch', 'Board',
            ])
            wb.save(excel_path)
        wb = load_workbook(excel_path)
        ws = wb.active
        # Prepare row
        row = [
            result.get('board_number'),
            result.get('timestamp'),
            result.get('result'),
            ', '.join(result.get('missing', [])),
            str(result.get('detected', {})),
            str(result.get('expected', {})),
            result.get('batch'),
            result.get('board'),
        ]
        ws.append(row)
        wb.save(excel_path)

    def export_batch(self, name, path):
        if name in self.batches:
            with open(path, 'w') as f:
                json.dump(self.batches[name], f, indent=2)
